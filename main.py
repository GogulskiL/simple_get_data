import requests
import bs4
from openpyxl import Workbook

url = 'https://rpa.hybrydoweit.pl'


def get_data(url):
    return requests.get(url)


def create_soup(url):
    data = get_data(url)
    return bs4.BeautifulSoup(data.text, 'html.parser')


def get_article_data(url):
    articles_section = create_soup(url).find(
        id="articles").find_all('article', class_='rpa-article-card')
    return [[(tag.findChild('a')['title']), ' '.join((tag.find('li').string.split()[
        1:])), (tag.findChild('a')['href'])] for tag in articles_section]


def get_article_data_reverse(url):
    return get_article_data(url)[::-1]


def create_file(name_file, url):
    title = "Tytuł"
    industry = "Branża/tytuł"
    link = "Link"
    data = get_article_data(url)
    data_rev = get_article_data_reverse(url)
    wb = Workbook()
    sheet = wb.active
    sheet_1 = wb.create_sheet("Sheet 1")

    sheet.cell(row=1, column=1).value = sheet_1.cell(
        row=1, column=1).value = title
    sheet.cell(row=1, column=2).value = sheet_1.cell(
        row=1, column=2).value = industry
    sheet.cell(row=1, column=3).value = sheet_1.cell(
        row=1, column=3).value = link

    for i in range(2, 7):
        for k in range(0, 3):
            sheet.cell(row=i, column=k+1).value = data[i-2][k]

    for i in range(2, 7):
        for k in range(0, 3):
            sheet_1.cell(row=i, column=k+1).value = data_rev[i-2][k]

    wb.save(f"{name_file}.xlsx")


if __name__ == '__main__':
    create_file("excel", url)
